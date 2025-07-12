from collections import defaultdict
from datetime import datetime
from typing import overload
from dateutil.relativedelta import relativedelta
import misc.ui as ui
import openpyxl
import os
from tqdm import tqdm
from venues.records import VenueRecord
from venues.errors import HashError, NoValidSessionsException
from openpyxl.styles import Border, Side, Alignment
from openpyxl.styles import Font

expected_headers = [
            'Job#', 'User', 'MKT', 'LOC#', 'Week', 'Zone', 'Restaurant',
            'St Address', 'City', 'ST', 'ZIP', 'Mail Piece', 'Month', 'Year',
            '# Sessions', 'Qty', 'RSVPs', 'RMI',
            
            'Lunch Day 1', 'Lunch 1 Date', 'Lunch 1 Time',
            'Lunch Day 2', 'Lunch 2 Date', 'Lunch 2 Time',
            'Lunch Day 3', 'Lunch 3 Date', 'Lunch 3 Time',
            'Dinner Day 1', 'Dinner 1 Date', 'Dinner 1 Time',
            'Dinner Day 2', 'Dinner 2 Date', 'Dinner 2 Time',
            'Dinner Day 3', 'Dinner 3 Date', 'Dinner 3 Time']

@overload
def generate() -> None: ...
@overload
def generate(venue_records: set['VenueRecord']) -> None: ...

def generate(venue_records: set['VenueRecord']=None):
    print('\n[Begin new report]')
    if (venue_records is None):
        # Display logotype intro
        ui.hideCursor()
        ui.prompt_user('\nThis program will now prompt you to select an Excel (.xlsx) file containing venue data. Press any key to continue.')
        ui.pause()

        # Prompt for excel file
        file_path = _get_file_path(test=False)

        # Load file
        headers, raw_data_sheet = _load_excel(file_path)
        
        cutoff_date = ui.query_date(
            'Data Set Cutoff Date (MM/DD/YY): ',
            default=datetime.now() - relativedelta(months=16))
        
        print('Extracting data. This may take a minute...')
        venue_records = _extract_data(headers, raw_data_sheet, cutoff_date)
        ui.print_success('Extraction complete.')


    # ----- QUERY USER FOR PARAMETERS -----
    # Query scheduling dates
    print('\nPlease enter the scheduling period for this report. Scheduling period does not support default values.')
    
    start_date = ui.query_date('Start date (MM/DD/YY): ')
    end_date = ui.query_date('End Date (MM/DD/YY): ')

    # We can't have the start date be after the end date.
    while start_date > end_date:
        ui.print_error('Scheduling period start date cannot be after end date. Please try again.')
        start_date = ui.query_date('Start date (MM/DD/YY): ')
        end_date = ui.query_date('End Date (MM/DD/YY): ')

    
    # A bunch of other queries
    print('\nFor default values on any of the following questions, continue without entering anything.')
    # Query saturation period
    saturation_period = ui.query_int('Zone Saturation Period (weeks): ', 16)
    # Query for the "around the same time period"
    prox_weeks = ui.query_int('Scheduling Period Lookback Margin (weeks): ', 2)
    # Query minimum RSVPs
    min_rsvps = ui.query_int('Minimum RSVPs: ', 16)
    # Query minimum ROR value
    min_ror = ui.query_float('Minimum ROR (%): ', 0)
    # Query venue cap
    num_venues = ui.query_int('Number of venues per market: ', 20)

    # Query specific markets
    print('\nFor specific markets, use market codes separated by spaces (e.g., "HOU PDX...")')
    markets = ui.query_user('Specific Markets: ').split(' ')

    # Exclude venues...
    # 1. Whose last RSVPs do not meet min_rsvps, and
    # 2. Who are in a zone which has had a job within the last four months

    print('Executing set exclusions...')
    # We want to exclude all zones that have had an event within four months
    filtered_data = _filter_data(venue_records, saturation_period, start_date, min_rsvps, min_ror)

    # Split venues into those who had a job around the same time last year, and those that didn't
    # sort the proximal venues by ROR
    # sort the non-proximal venues first by oldest last job month
    # and then by ROR
    print('Performing optimizations...')
    proximal_venues = {
        venue for venue in filtered_data
        if venue.around_time_last_year(start_date, end_date, prox_weeks)
    }

    proximal_venues = sorted(proximal_venues, key=lambda venue: venue.latest_job.ror, reverse=True)

    nonproximal_venues = {
        venue for venue in filtered_data
        if not venue.around_time_last_year(start_date, end_date, prox_weeks)
    }

    nonproximal_venues = sorted(nonproximal_venues, key=lambda venue: venue.latest_job.ror, reverse=True)
    nonproximal_venues = sorted(nonproximal_venues, key=lambda venue: venue.latest_job.month_date)

    # Recombine so that proximal venues are on top.
    sorted_data = proximal_venues + nonproximal_venues


    ui.print_success('Exclusions and optimizations complete.')

    # Prepare to output data
    ui.prompt_user('\nThis program will now prompt you to select an ouput directory. Press any key to continue.')
    ui.pause()
    
    selected_dir = ui.promptDirectory()
    #selected_dir = 'C:\\Users\\alexc\\Documents\\data-direct\\test'

    if selected_dir == '':
        ui.print_error('No directory selected. Terminating report.')
        generate(venue_records)
        return

    print('Creating output directory...')
    output_dir = selected_dir + f'\\VEN_REPORT_{start_date.strftime("%m_%d_%y")}-{end_date.strftime("%m_%d_%y")}'

    # Check if directory already exists, and if so warn user. 
    try:
        os.makedirs(output_dir, exist_ok=False)
    except OSError:
        ui.print_warning('WARNING: A venues report folder with the same name already exists at the selected location. Please move it or select a different directory. This report will terminate.')
        generate(venue_records)
        return

    print('Classifying records by market...')
    venues_by_market = defaultdict(list[VenueRecord])
    for venue in sorted_data:
        venues_by_market[venue.market].append(venue)
    
    print('Writing records to new files...')
    # Write to new excel file
    for market, venues in venues_by_market.items():
        # If user requested specific markets, halt for non-specified markets
        if markets[0] != '' and market not in markets:
            continue

        wb = openpyxl.Workbook()
        ws = wb.active

        # IMPORTANT if headers are changed, then the returned tuple
        # from Venue.to_entry() must also be changed so that the header
        # order matches the data order.
        headers = [
            'Job#', 'User', 'MKT', 'LOC#', 'Week', 'Zone', 'Zone/Last',
            'Restaurant', 'St Address', 'City', 'ST', 'ZIP',
            'Mail Piece', 'Qty', 'Venue/Last', '# Sessions', 
            'Session Type', 'RSVPs', 'RMI', 'ROR%', 'Venue/Qualifier', 
            'RSVPs', 'ROR', 'Zone use within 12 months', 'Average ROR%']

        ws.append(headers)

        # This is for capping number of written venues
        i = 0

        for venue in venues:
            if i < num_venues:
                row = venue.to_entry(start_date, end_date, prox_weeks, venue_records)
                ws.append(row)
                i += 1

        _style_workbook(wb)

        file_path = os.path.join(output_dir, f'{market}_{start_date.strftime("%m_%d_%y")}-{end_date.strftime("%m_%d_%y")}.xlsx')

        wb.save(file_path)

    ui.print_success(f"Report(s) have been saved. Press any key to begin a new report, or close the program.")
    ui.pause()
    
    generate(venue_records)
    return



# ===== internal helper functions ===== #



def _get_file_path(test: bool=False) -> str:
    """Query the user for a filepath and returns a string. Will cause a UI
    error and exit if the user enters an empty string.
    """
    if test:
        file_path = 'C:\\Users\\alexc\\Documents\\data-direct\\test\\test_input.xlsx'
    else:
        file_path = ui.promptFile((('Excel Spreadsheet', ('*.xlsx')),('All files', '*.*')))


    # Validate file path
    if file_path == '':
        ui.print_error('No file was selected.')
        ui.pause()
        ui.exit()
    
    return file_path


def _load_excel(excel_file_path: str) -> list:
    """Attempt to load an excel spreadsheet and return a list which
    can be iterated over. Will cause a UI error if an exception occurs or
    the file is missing necessary file headers, which are hardcoded in this
    class.
    """
    try:
        # Read file with openpyxl
        print('Loading Excel file...')
        workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
       
        missing_headers = [exp_hdr for exp_hdr in expected_headers if exp_hdr not in headers]
        # If there are missing headers
        if len(missing_headers) > 0:
            missing_headers_msg = 'The selected file is missing the following expected columns:'
            for header in missing_headers:
                missing_headers_msg += f'\n{header}'
            ui.print_error(missing_headers_msg)
            ui.pause()
            ui.exit()
        else:
            ui.print_success('All expected headers are present.')

    except BaseException as e:
        ui.print_error(f'An error occured while reading the file. This is likely due to invalid file format.')
        ui.pause()
        ui.exit()
    
    return (headers, sheet)


def _extract_data(headers: list[str], raw_data_sheet: list, cutoff_date: datetime) -> set['VenueRecord']:
    """Accepts a data sheet like one from an openpyxl workbook and retrns
    a set of VenueRecords. Will skip over malformed entries in the sheet
    without raising any exceptions.
    """
    venue_records: set[VenueRecord] = set()
    # Load data into structures
    # Iterate through each entry
    for entry in tqdm(raw_data_sheet.iter_rows(min_row=2, values_only=True), total=raw_data_sheet.max_row - 1):
        # If entry contains a date before cutoff date, don't evaluate
        outdated = False
        for val in entry:
            try:
                if val < cutoff_date:
                    outdated = True
                    break
            except:
                pass
        if outdated:
            continue

        # Convert tuple to dict so we can reference by key
        entry = dict(zip(headers, entry))

        # If no job id, then there's not really an entry here
        if entry['Job#'] is None:
            continue

        # Create a new venue (or at least try to)
        try:
            new_venue = VenueRecord.from_entry(entry)
            found = False
            # And check if it matches an existing venue
            for existing_venue in venue_records:
                # If there is a match
                if new_venue == existing_venue:
                    # Add new job record to existing venue
                    existing_venue.add_job_record(entry)  
                    found = True

            # If no matching venue found, add this one to the set
            if not found:
                venue_records.add(new_venue)

        except NoValidSessionsException:
            #tqdm.write(ui.warning(f'No valid sessions found for job {entry['Job#']}. Skipping this job.'))
            pass

        except (TypeError, ValueError) as e:
            #if entry['Job#'] is not None:
            #    tqdm.write(ui.warning(f'Job {entry['Job#']} is invalidly formatted. Skipping this job.'))
            pass

        except (HashError) as e:
            pass

        #except BaseException:
        #    if entry['Job#'] is not None:
        #        #ui.print_warning(f'Job {entry['Job#']} is invalidly formatted. Skipping job.')
        #        # TODO - printing a warning is too verbose. Maybe do something else?
        #        pass

    return venue_records


def _filter_data(venue_records: set[VenueRecord], saturation_period: int, start_date: datetime, min_rsvps: int, min_ror: float):
    """Filters out undesirable venues. The current criteria is based on minimum
    number of RSVPs and whether a venue's zone has had a seminar within the `saturation_period` (weeks).
    """
    
    saturated_zones = {
        venue.zone for venue in venue_records 
        if venue.jobs_within(relativedelta(weeks=saturation_period), start_date)
    }

    # Filter by saturated zones and minimum rsvps
    filtered_data = {
        venue for venue in venue_records
        if (venue.zone not in saturated_zones
            and venue.latest_job.rvsps >= min_rsvps
            and venue.latest_job.ror >= min_ror)
    }

    return filtered_data


def _style_workbook(wb: openpyxl.Workbook):
    """Add styles to the workbook.
    """

    # Make headers bold and pinned (freeze top row)
    for ws in wb.worksheets:
        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
        # Freeze top row
        ws.freeze_panes = ws['A2']

    # Left justify, pad, and border cells
    left_alignment = Alignment(horizontal="left", vertical="center", indent=0, wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        col_maxlen = {}
        
        # Skip empty worksheets
        if ws.max_row == 0:
            continue
            
        for row_num, row in enumerate(ws.iter_rows(), start=1):
            for idx, cell in enumerate(row, start=1):
                # Center and pad
                cell.alignment = left_alignment
                # Add black border
                cell.border = border
                
                # Track max length for column width, excluding headers (row 1)
                if row_num > 1:  # Skip header row for width calculation
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    col_maxlen[idx] = max(col_maxlen.get(idx, 0), cell_len)

        # Set column widths based on data rows only
        for idx, maxlen in col_maxlen.items():
            col_letter = openpyxl.utils.get_column_letter(idx)
            # Add padding and set a reasonable minimum width
            width = max(maxlen + 4, 8)  # Minimum width of 8
            ws.column_dimensions[col_letter].width = width
